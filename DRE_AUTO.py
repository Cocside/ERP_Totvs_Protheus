"""
DRE Report Processor
====================
Monitora a pasta DRE, processa arquivos .xlsx com DRE no nome, criando uma Pivot Table nativa,
move o arquivo processado e envia por e-mail.

@author  Helder Vieira Medeiros
@since   28/04/2026
"""

import os
import re
import time
import shutil
import logging
import smtplib
import zipfile
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ---------------------------------------------------------------------------
# Configuração de pastas
# ---------------------------------------------------------------------------
WATCH_DIR   = Path(r"E:\totvs\protheus12\ambientes\producao\Protheus_Data\integracao_smartview\DRE")
OUTPUT_DIR  = Path(r"E:\totvs\protheus12\ambientes\producao\Protheus_Data\integracao_smartview\pronto")
BACKUP_DIR  = Path(r"E:\totvs\protheus12\ambientes\producao\Protheus_Data\integracao_smartview\backup")
TARGET_FILE = "DRE_VISAO.xlsx"

# ---------------------------------------------------------------------------
# Configuração de e-mail
# ---------------------------------------------------------------------------
SMTP_SERVER   = os.getenv("SMTP_SERVER",   "smtp.office365.com")
SMTP_PORT     = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER     = os.getenv("SMTP_USER",     "totvs.notificacao@jallcard.com.br")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "9a5;2uNP&h0C|0!@#")  # <-- use variável de ambiente

EMAIL_TO      = "totvs.resultados@jallcard.com.br"
#EMAIL_TO      = "helder.medeiros@jallcard.com.br"
EMAIL_SUBJECT = "Relatório semanal DRE"
EMAIL_BODY    = "Prezados,\n\nSegue a DRE do período.\n\nMensagem automática. Para dúvidas ou ajustes, por favor entrar em contato com o suporte.\n\nAtenciosamente,\n\nSuporte Totvs,"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),

        logging.FileHandler("dre_processor.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers de parsing XML/OOXML  (tolerantes à ordem dos atributos)
# ---------------------------------------------------------------------------

def _parse_workbook_sheets(wb_xml: str) -> list[dict]:
    """
    Retorna lista de dicts {name, sheetId, rid} usando xml.etree.
    Tem fallback em regex que cobre qualquer ordem de atributos.
    """
    import xml.etree.ElementTree as ET

    r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    try:
        root = ET.fromstring(wb_xml)
        sheets = []
        for elem in root.iter(f"{{{main_ns}}}sheet"):
            sheets.append({
                "name":    elem.get("name"),
                "sheetId": elem.get("sheetId"),
                "rid":     elem.get(f"{{{r_ns}}}id"),
            })
        if sheets:
            return sheets
    except ET.ParseError:
        pass

    # Fallback: regex que aceita atributos em qualquer ordem dentro da tag <sheet ...>
    sheets = []
    for tag in re.findall(r"<sheet\b([^/?>]+)", wb_xml):
        name    = re.search(r'\bname=["\']([^"\']+)["\']', tag)
        sheet_id = re.search(r'\bsheetId=["\']([^"\']+)["\']', tag)
        rid     = re.search(r'\br:id=["\']([^"\']+)["\']', tag)
        if name and rid:
            sheets.append({
                "name":    name.group(1),
                "sheetId": sheet_id.group(1) if sheet_id else None,
                "rid":     rid.group(1),
            })
    return sheets


def _parse_rels(rels_xml: str) -> dict[str, str]:
    """Retorna {Id -> Target} de um arquivo .rels."""
    import xml.etree.ElementTree as ET

    mapping = {}
    try:
        root = ET.fromstring(rels_xml)
        for elem in root.iter():
            rid    = elem.get("Id")
            target = elem.get("Target")
            if rid and target:
                mapping[rid] = target
        if mapping:
            return mapping
    except ET.ParseError:
        pass

    # Fallback regex — aceita qualquer ordem de Id / Target
    for tag in re.findall(r"<Relationship\b([^/?>]+)", rels_xml):
        rid    = re.search(r'\bId=["\']([^"\']+)["\']', tag)
        target = re.search(r'\bTarget=["\']([^"\']+)["\']', tag)
        if rid and target:
            mapping[rid.group(1)] = target.group(1)
    return mapping


# ---------------------------------------------------------------------------
# Construção da Pivot Table (XML OOXML)
# ---------------------------------------------------------------------------

def _xml_escape(value: str) -> str:
    return (
        value
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _build_cache_definition(data_sheet: str, df: pd.DataFrame) -> str:
    nrows    = len(df) + 1
    last_col = get_column_letter(len(df.columns))

    fields_xml = ""
    for col in df.columns:
        unique_vals = sorted(df[col].dropna().unique(), key=str)
        count       = len(unique_vals)
        if col == "VALOR":
            fields_xml += (
                f'<cacheField name="{_xml_escape(col)}" numFmtId="4">'
                f'<sharedItems containsNonDate="1" containsDate="0" containsString="0" '
                f'containsNumber="1" containsInteger="0" count="{count}"/>'
                f'</cacheField>'
            )
        else:
            items = "".join(f'<s v="{_xml_escape(str(v))}" />' for v in unique_vals)
            fields_xml += (
                f'<cacheField name="{_xml_escape(col)}" numFmtId="0">'
                f'<sharedItems count="{count}">{items}</sharedItems>'
                f'</cacheField>'
            )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotCacheDefinition'
        ' xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' r:id="rIdCache1" refreshedBy="Python" refreshedDate="0"'
        ' createdVersion="4" refreshedVersion="4" minRefreshableVersion="3"'
        ' recordCount="0" refreshOnLoad="1">'
        '<cacheSource type="worksheet">'
        f'<worksheetSource ref="A1:{last_col}{nrows}" sheet="{_xml_escape(data_sheet)}"/>'
        '</cacheSource>'
        f'<cacheFields count="{len(df.columns)}">{fields_xml}</cacheFields>'
        '</pivotCacheDefinition>'
    )


def _build_pivot_definition(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    fi   = {c: i for i, c in enumerate(cols)}
    n    = len(cols)

    pf_list = []
    for col in cols:
        if col == "ANO":
            pf_list.append(
                '<pivotField name="ANO" axis="axisPage" showAll="0">'
                '<items count="1"><item t="default"/></items></pivotField>'
            )
        elif col == "MES":
            pf_list.append(
                '<pivotField name="MES" axis="axisCol" showAll="0">'
                '<items count="1"><item t="default"/></items></pivotField>'
            )
        elif col == "LINHA_DRE":
            pf_list.append(
                '<pivotField name="LINHA_DRE" axis="axisRow" showAll="0" outline="1" subtotalTop="1">'
                '<items count="1"><item t="default"/></items></pivotField>'
            )
        elif col == "CONTA_DETALHE":
            pf_list.append(
                '<pivotField name="CONTA_DETALHE" axis="axisRow" showAll="0" outline="1" subtotalTop="0">'
                '<items count="1"><item t="default"/></items></pivotField>'
            )
        elif col == "VALOR":
            pf_list.append('<pivotField name="VALOR" dataField="1" showAll="0"/>')
        else:
            pf_list.append(f'<pivotField name="{_xml_escape(col)}" showAll="0"/>')

    f_linha = fi["LINHA_DRE"]
    f_conta = fi["CONTA_DETALHE"]
    f_mes   = fi["MES"]
    f_ano   = fi["ANO"]
    f_valor = fi["VALOR"]

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotTableDefinition'
        ' xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' name="DRE_JALLCARD" cacheId="1" dataCaption="Valores"'
        ' applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0"'
        ' applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"'
        ' showMemberPropertyTips="0" useAutoFormatting="1" itemPrintTitles="1"'
        ' createdVersion="4" indent="0" outline="1" outlineData="1"'
        ' multipleFieldFilters="0" showDrill="1" showDataDropDown="1">'
        '<location ref="A3" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"'
        ' rowPageCount="1" colPageCount="1"/>'
        f'<pivotFields count="{n}">{"".join(pf_list)}</pivotFields>'
        f'<rowFields count="2"><field x="{f_linha}"/><field x="{f_conta}"/></rowFields>'
        f'<colFields count="1"><field x="{f_mes}"/></colFields>'
        f'<pageFields count="1"><pageField fld="{f_ano}" hier="-1"/></pageFields>'
        f'<dataFields count="1">'
        f'<dataField name="Soma de VALOR" fld="{f_valor}" subtotal="sum" showDataAs="normal" numFmtId="4"/>'
        f'</dataFields>'
        '</pivotTableDefinition>'
    )


# ---------------------------------------------------------------------------
# Injeção da Pivot Table no pacote OOXML
# ---------------------------------------------------------------------------

def _inject_pivot(xlsx_path: Path, data_sheet: str, pivot_sheet: str, df: pd.DataFrame):
    pivot_def_xml = _build_pivot_definition(df)
    cache_def_xml = _build_cache_definition(data_sheet, df)
    tmp_path      = xlsx_path.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:

        names = zin.namelist()

        # --- Descobre o arquivo .xml da aba pivot ---
        wb_xml      = zin.read("xl/workbook.xml").decode("utf-8")
        wb_rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        sheets   = _parse_workbook_sheets(wb_xml)
        rels_map = _parse_rels(wb_rels_xml)

        log.debug(f"Sheets detectadas: {sheets}")
        log.debug(f"Rels detectadas: {rels_map}")

        pivot_rid = None
        for s in sheets:
            if s["name"] == pivot_sheet:
                pivot_rid = s["rid"]
                break

        if pivot_rid is None:
            raise ValueError(
                f"Aba '{pivot_sheet}' não encontrada. "
                f"Sheets disponíveis: {[s['name'] for s in sheets]}"
            )

        pivot_ws_file = rels_map.get(pivot_rid)
        if pivot_ws_file is None:
            raise ValueError(
                f"rId '{pivot_rid}' não encontrado nas rels. "
                f"Rels disponíveis: {list(rels_map.keys())}"
            )

        m = re.search(r"sheet(\d+)", pivot_ws_file)
        if m is None:
            raise ValueError(f"Não foi possível extrair número da sheet de '{pivot_ws_file}'")
        pivot_ws_num = m.group(1)

        pivot_ws_rels_path    = f"xl/worksheets/_rels/sheet{pivot_ws_num}.xml.rels"
        pivot_cache_path      = "xl/pivotCache/pivotCacheDefinition1.xml"
        pivot_records_path    = "xl/pivotCache/pivotCacheRecords1.xml"
        pivot_table_path      = "xl/pivotTables/pivotTable1.xml"
        pivot_table_rels_path = "xl/pivotTables/_rels/pivotTable1.xml.rels"

        # --- Copia e patcha arquivos existentes ---
        for item in names:
            data = zin.read(item)

            if item == "[Content_Types].xml":
                content = data.decode("utf-8")
                if "pivotTable" not in content:
                    content = content.replace(
                        "</Types>",
                        '<Override PartName="/xl/pivotTables/pivotTable1.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>'
                        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
                        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
                        "</Types>",
                    )
                zout.writestr(item, content.encode("utf-8"))

            elif item == "xl/_rels/workbook.xml.rels":
                content = data.decode("utf-8")
                if "pivotCacheDefinition" not in content:
                    content = content.replace(
                        "</Relationships>",
                        '<Relationship Id="rIdPivotCache1"'
                        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"'
                        ' Target="pivotCache/pivotCacheDefinition1.xml"/>'
                        "</Relationships>",
                    )
                zout.writestr(item, content.encode("utf-8"))

            elif item == "xl/workbook.xml":
                content = data.decode("utf-8")
                if "pivotCaches" not in content:
                    content = content.replace(
                        "</workbook>",
                        '<pivotCaches>'
                        '<pivotCache cacheId="1" r:id="rIdPivotCache1"/>'
                        '</pivotCaches></workbook>',
                    )
                zout.writestr(item, content.encode("utf-8"))

            else:
                zout.writestr(item, data)

        # --- Escreve novos arquivos da Pivot ---
        zout.writestr(
            pivot_ws_rels_path,
            (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdPivot1"'
                ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"'
                ' Target="../pivotTables/pivotTable1.xml"/>'
                '</Relationships>'
            ).encode("utf-8"),
        )
        zout.writestr(
            pivot_table_rels_path,
            (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdCache1"'
                ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"'
                ' Target="../pivotCache/pivotCacheDefinition1.xml"/>'
                '</Relationships>'
            ).encode("utf-8"),
        )
        zout.writestr(pivot_table_path,  pivot_def_xml.encode("utf-8"))
        zout.writestr(pivot_cache_path,  cache_def_xml.encode("utf-8"))
        zout.writestr(
            pivot_records_path,
            (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<pivotCacheRecords'
                ' xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'
            ).encode("utf-8"),
        )

    tmp_path.replace(xlsx_path)
    log.info("Pivot Table injetada com sucesso.")



# ---------------------------------------------------------------------------
# Processamento principal
# ---------------------------------------------------------------------------

def process_file(source_path: Path) -> Path | None:
    try:
        log.info(f"Lendo: {source_path}")
        df = pd.read_excel(source_path, header=1, dtype=str)
        df.columns = [c.strip().upper() for c in df.columns]

        required = {"ANO", "MES", "LINHA_DRE", "CONTA_DETALHE", "VALOR"}
        missing  = required - set(df.columns)
        if missing:
            log.error(f"Colunas ausentes: {missing}")
            return None

        df["VALOR"] = (
            pd.to_numeric(df["VALOR"].str.replace(",", "."), errors="coerce").fillna(0)
        )

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        ts          = datetime.now().strftime("%d-%m-%Y")
        output_path = OUTPUT_DIR / f"DRE_JALLCARD_{ts}.xlsx"

        wb = openpyxl.Workbook()

        # Aba Dados
        ws_data       = wb.active
        ws_data.title = "Dados"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        thin        = Side(style="thin", color="D9D9D9")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, col in enumerate(df.columns, start=1):
            cell           = ws_data.cell(row=1, column=ci, value=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        data_font = Font(name="Arial", size=10)
        alt_fill  = PatternFill("solid", fgColor="EBF3FB")
        for ri, row in enumerate(df.itertuples(index=False), start=2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row, start=1):
                cell        = ws_data.cell(row=ri, column=ci, value=val)
                cell.font   = data_font
                cell.border = border
                if fill:
                    cell.fill = fill
                if df.columns[ci - 1] == "VALOR":
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")

        for ci, col in enumerate(df.columns, start=1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max()) + 4
            ws_data.column_dimensions[get_column_letter(ci)].width = min(max_len, 40)

        ws_data.freeze_panes = "A2"
        ws_data.sheet_state  = "hidden"  # <-- Oculta a aba Dados no Excel final

        nrows = len(df) + 1
        tbl   = Table(
            displayName="TabelaDRE",
            ref=f"A1:{get_column_letter(len(df.columns))}{nrows}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            ),
        )
        ws_data.add_table(tbl)

        # Aba Pivot (placeholder visual — XML real é injetado depois)
        ws_pivot       = wb.create_sheet("DRE_JALLCARD")
        ws_pivot["A1"] = "Tabela Dinâmica – DRE"
        ws_pivot["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Arial")
        ws_pivot["A2"] = "Filtro: ANO | Colunas: MÊS | Linhas: LINHA_DRE > CONTA_DETALHE | Valores: VALOR"
        ws_pivot["A2"].font = Font(italic=True, size=9, color="595959", name="Arial")

        wb.save(output_path)
        log.info(f"Arquivo base salvo: {output_path}")

        _inject_pivot(output_path, ws_data.title, ws_pivot.title, df)

        log.info(f"Processamento concluído: {output_path}")
        return output_path

    except PermissionError:
        log.warning("Arquivo bloqueado por outro processo.")
        return None
    except Exception:
        log.exception("Erro ao processar arquivo.")
        return None


# ---------------------------------------------------------------------------
# Envio de e-mail
# ---------------------------------------------------------------------------

def send_email(attachment_path: Path):
    try:
        msg            = MIMEMultipart()
        msg["From"]    = SMTP_USER
        msg["To"]      = EMAIL_TO
        msg["Subject"] = EMAIL_SUBJECT
        msg.attach(MIMEText(EMAIL_BODY, "plain"))

        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{attachment_path.name}"',
        )
        msg.attach(part)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_USER, EMAIL_TO, msg.as_string())

        log.info(f"E-mail enviado para {EMAIL_TO} | anexo: {attachment_path.name}")

    except Exception:
        log.exception("Falha ao enviar e-mail.")


# ---------------------------------------------------------------------------
# Watchdog Handler  (com debounce para evitar double-trigger)
# ---------------------------------------------------------------------------

class DREHandler(FileSystemEventHandler):
    def __init__(self):
        self._processing   = False
        self._last_trigger = 0.0
        self._DEBOUNCE_S   = 5  # segundos de cooldown

    def on_created(self, event):
        self._check(event.src_path)

    def on_moved(self, event):
        self._check(event.dest_path)

    def on_modified(self, event):
        self._check(event.src_path)

    def _check(self, path: str):
        p = Path(path)
        now = time.time()

        # NOVA LÓGICA DE FILTRO:
        # 1. Verifica se é um arquivo (ignora pastas)
        # 2. Verifica se a extensão é .xlsx
        # 3. Verifica se "DRE" está no nome (case-insensitive)
        if not p.is_file() or p.suffix.lower() != ".xlsx" or "DRE" not in p.name.upper():
            return

        # Ignora arquivos temporários do Excel (que começam com ~$)
        if p.name.startswith("~$"):
            return

        if self._processing or (now - self._last_trigger) < self._DEBOUNCE_S:
            log.debug(f"Evento ignorado (debounce/processando): {p.name}")
            return

        self._processing = True
        self._last_trigger = now
        log.info(f"Arquivo detectado e válido para processamento: {p.name}")

        # O restante do código de espera e processamento permanece igual...
        # Aguarda arquivo ser liberado
        for attempt in range(10):
            try:
                with open(p, "rb"):
                    break
            except (PermissionError, OSError):
                log.warning(f"Arquivo {p.name} bloqueado, aguardando... ({attempt + 1}/10)")
                time.sleep(3)
        else:
            log.error(f"Arquivo {p.name} permanentemente bloqueado. Abortando.")
            self._processing = False
            return

        output_path = process_file(p)

        if output_path:
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%d-%m-%Y")
            # Ajuste no nome do backup para manter o nome original do arquivo
            backup_path = BACKUP_DIR / f"{p.stem}_backup_{ts}{p.suffix}"
            shutil.move(str(p), str(backup_path))
            log.info(f"Original movido para backup: {backup_path}")
            send_email(output_path)

        self._processing = False


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def main():
    WATCH_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    handler  = DREHandler()
    observer = Observer()
    observer.schedule(handler, str(WATCH_DIR), recursive=False)
    observer.start()
    log.info(f"Monitorando: {WATCH_DIR}")
    log.info("Pressione Ctrl+C para encerrar.")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
    log.info("Serviço encerrado.")


if __name__ == "__main__":
    main()